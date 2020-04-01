#!/usr/bin/env python
#
# Copyright 2015, 2016 Thomas Timm Andersen (original version)
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import time
import math
import roslib; roslib.load_manifest('ur_driver')
import rospy
import actionlib
from openpyxl import Workbook
from control_msgs.msg import *
from trajectory_msgs.msg import *
from sensor_msgs.msg import JointState
from math import pi
import numpy as np
import tensorflow as tf
import numpy as np
from scipy.integrate import odeint
from math import exp

# From Files
from object_detection import IntelRealsense
from universal_robot_kinematics import invKine
from kinematics import fwd_kin
from last_kalman_filter import *

IntelRealsense = IntelRealsense()

JOINT_NAMES = ['shoulder_pan_joint', 'shoulder_lift_joint', 'elbow_joint',
               'wrist_1_joint', 'wrist_2_joint', 'wrist_3_joint']
home = [0, -pi/2, pi/2, 0, pi/2, pi]
client = None

# Kanut Thummaraksa
#### Input Tensors ####
 ## Common Input ##
s = tf.placeholder(tf.float64,name='s')
tau = tf.placeholder(tf.float64,name='tau')

xg = tf.placeholder(tf.float64,name='xg')
yg = tf.placeholder(tf.float64,name='yg')
zg = tf.placeholder(tf.float64,name='zg')

 ## joints ##
g = (tf.placeholder(tf.float64,name='g1'),
    tf.placeholder(tf.float64,name='g2'),
    tf.placeholder(tf.float64,name='g3'),
    tf.placeholder(tf.float64,name='g4'),
    tf.placeholder(tf.float64,name='g5'),
    tf.placeholder(tf.float64,name='g6'))

q = (tf.placeholder(tf.float64,name='q1'),
    tf.placeholder(tf.float64,name='q2'),
    tf.placeholder(tf.float64,name='q3'),
    tf.placeholder(tf.float64,name='q4'),
    tf.placeholder(tf.float64,name='q5'),
    tf.placeholder(tf.float64,name='q6'))

qd = (tf.placeholder(tf.float64,name='qd1'),
    tf.placeholder(tf.float64,name='qd2'),
    tf.placeholder(tf.float64,name='qd3'),
    tf.placeholder(tf.float64,name='qd4'),
    tf.placeholder(tf.float64,name='qd5'),
    tf.placeholder(tf.float64,name='q06'))

q0 = (tf.placeholder(tf.float64,name='q01'),
    tf.placeholder(tf.float64,name='q02'),
    tf.placeholder(tf.float64,name='q03'),
    tf.placeholder(tf.float64,name='q04'),
    tf.placeholder(tf.float64,name='q05'),
    tf.placeholder(tf.float64,name='q06'))

def canoSystem(tau,t):
    alpha_s = 4
    s = exp(-tau*alpha_s*t)
    return s

def dmp(g,q,qd,tau,s,q0,W,Name = "DMP"):
    alpha = tf.constant(25,dtype=tf.float64)
    beta = alpha/4
    w,c,h = W
    n_gaussian = w.shape[0]
    with tf.name_scope(Name):
        w_tensor = tf.constant(w,dtype=tf.float64,name='w')
        c_tensor = tf.constant(c,dtype=tf.float64,name='c')
        h_tensor = tf.constant(h,dtype=tf.float64,name='h')

        with tf.name_scope('s'):
            s_tensor = s*tf.ones(n_gaussian,dtype=tf.float64)
        smc_pow = tf.pow(s_tensor-c_tensor,2)
        h_smc_pow = tf.math.multiply(smc_pow,(-h_tensor))
        with tf.name_scope('psi'):
            psi = tf.math.exp(h_smc_pow)
        sum_psi = tf.math.reduce_sum(psi,0)
        wpsi = tf.math.multiply(w_tensor,psi)
        wpsis = tf.math.reduce_sum(wpsi*s,0)
        with tf.name_scope('fs'):
            fs =wpsis/sum_psi
        qdd = alpha*(beta*(g-q)-tau*qd)+fs*(g-q0)
    return qdd

#### Movement Library #####
dmps = [{},{},{},{},{},{}]
for i in range(15):
 path = 'Demonstration/Demo{}/Weights/'.format(i+1)
 for j in range(6): ### j = joint number
     path_j = path+'Joint{}/'.format(j+1)
     w = np.load(path_j+'w.npy')
     w = np.reshape(w,(len(w),))
     c = np.load(path_j+'c.npy')
     c = np.reshape(c,(len(c),))
     h = np.load(path_j+'h.npy')
     h = np.reshape(h,(len(h),))
     W = (w,c,h)
     # def dmp(g,q,qd,tau,s,q0,W,Name = "DMP"):
     dmps[j]['{}_{}'.format(j+1,i+1)] =tf.reshape(dmp(g[j], q[j], qd[j], tau, s, q0[j], W, Name="DMP{}_{}".format(j+1,i+1)),(1,))

##### Final Catesian Position of Demonstration) #####
demo_x = np.array([-8.15926729e-01, -0.75961731, -0.3964087, -0.29553788, -0.04094927, -0.14693912, -0.41827111, -8.16843140e-01, -0.09284764, -0.57153495, -0.67251442, -0.36517125, -7.62308039e-01, -0.78029185, -6.57512038e-01])
demo_y = np.array([-2.96043917e-01, -0.18374539, 0.6690932,  0.21733157,  0.78624892,  0.7281835,   -0.66857267, -2.92201916e-01, -0.77947085, -0.28442803, 0.36890422,  -0.41997883, -1.20031233e-01, -0.19321253, -1.05877890e-01])
demo_z = np.array([-3.97988321e-03, 0.35300285,  0.13734106, 0.1860831,   0.06178831,  0.06178831,  0.10958549,  -5.64177448e-03, 0.0383235,   0.33788756,  0.30410704,  0.47738503,  8.29937352e-03,  0.17253172,  3.62063583e-01])


#### Contributin Functions ####
with tf.name_scope("Con"):
    xg_ref = tf.constant(demo_x, dtype=tf.float64,name="x_con")
    yg_ref = tf.constant(demo_y, dtype=tf.float64,name="y_con")
    zg_ref = tf.constant(demo_z, dtype=tf.float64,name="z_con")

    xg2 = tf.pow(xg_ref-xg, 2)
    yg2 = tf.pow(yg_ref-yg, 2)
    zg2 = tf.pow(zg_ref-zg, 2)

    sum = xg2+yg2+zg2
    con = 1.9947114020071635 * tf.math.exp(-0.5*sum/0.4472135954999579) # Normal Distribution


#### Gating Network #####
dmp_joint = []
dmpNet = []
for i in range(len(dmps)):
    values = list(dmps[i].values())
    joint = tf.concat(values, axis=0)
    with tf.name_scope('DMPNet{}'.format(i+1)):
        dmpNet_i = tf.reduce_sum(tf.math.multiply(joint,con),axis=0)/tf.reduce_sum(con, axis=0)
    dmpNet.append(dmpNet_i)

# Supitcha Klanpradit
def move_dmp_path(path_from_ode,time_from_ode):
    g = FollowJointTrajectoryGoal()
    g.trajectory = JointTrajectory()
    g.trajectory.joint_names = JOINT_NAMES
    try:
        for via in range(0,len(path_from_ode)):
            joint_update = path_from_ode[via][0:6]
            joint_update[0:5] = joint_update[0:5] - (joint_update[0:5]>math.pi)*2*math.pi + (joint_update[0:5]<-math.pi)*2*math.pi
            # print('Step %d %s' % (via,joint_update))
            g.trajectory.points.append(JointTrajectoryPoint(positions=joint_update, velocities=[0]*6, time_from_start=rospy.Duration(time_from_ode[via])))
        client.send_goal(g)
        client.wait_for_result()
    except KeyboardInterrupt:
        client.cancel_goal()
        raise
    except:
        raise

def set_home(set_position=home, set_duration=10):
    g = FollowJointTrajectoryGoal()
    g.trajectory = JointTrajectory()
    g.trajectory.joint_names = JOINT_NAMES
    try:
        g.trajectory.points = [JointTrajectoryPoint(positions=set_position, velocities=[0]*6, time_from_start=rospy.Duration(set_duration))]
        client.send_goal(g)
        client.wait_for_result()
    except KeyboardInterrupt:
        client.cancel_goal()
        raise
    except:
        raise

def cost_func(out_invKine):
    out_invKine[0:5,:] = out_invKine[0:5,:] - (out_invKine[0:5,:]>math.pi)*2*math.pi + (out_invKine[0:5,:]<-math.pi)*2*math.pi
    # print('inverse pingpong %s' %out_invKine)
    weight = [1, 1.2, 1.2, 1, 1, 1]
    weight = np.resize(weight,(6,8))
    cost = np.multiply(np.square(out_invKine), weight)
    cost = np.sum(cost, axis=0)
    # print('cost %s' %cost)
    index_minimum = np.argmin(cost)
    print('index minimum %s' %index_minimum)
    return [joint[0,index_minimum] for joint in out_invKine]

def choose_q(out_invKine, index):
    out_invKine[0:5,:] = out_invKine[0:5,:] - (out_invKine[0:5,:]>math.pi)*2*math.pi + (out_invKine[0:5,:]<-math.pi)*2*math.pi
    # print(out_invKine)1)
    for i in range(8):
        q = [joint[0,i] for joint in out_invKine]
        print('q %s', q)
        print('fwd %s' %fwd_kin(q, o_unit = 'p'))
    return [joint[0,index] for joint in out_invKine]

mu = 0
sigma_a_x = 0.8
sigma_a_y = 0.8
sigma_a_d = 0.8
rho_x = 0.1
rho_y = 0.1
rho_d = 0.1
rho_v_x = 0.5
rho_v_y = 0.5
rho_v_d = 0.5

def main():
    with tf.Session() as sess:
        sess.run(tf.global_variables_initializer())

        def dynamics(x,t,tau_v,g_v,q0_v,grasping_point,sess):
            position = grasping_point
            s_v = canoSystem(tau_v,t)
            feeddict = {g[0]:g_v[0],g[1]:g_v[1],g[2]:g_v[2],g[3]:g_v[3],g[4]:g_v[4],g[5]:g_v[5],
                        q[0]:x[0],q[1]:x[1],q[2]:x[2],q[3]:x[3],q[4]:x[4],q[5]:x[5],
                        qd[0]:x[6],qd[1]:x[7],qd[2]:x[8],qd[3]:x[9],qd[4]:x[10],qd[5]:x[11],
                        q0[0]:q0_v[0],q0[1]:q0_v[1],q0[2]:q0_v[2],q0[3]:q0_v[3],q0[4]:q0_v[4],q0[5]:q0_v[5],
                        tau:tau_v,s:s_v,xg:position[0],yg:position[1],zg:position[2]
                        }
            qdd1_v,qdd2_v,qdd3_v,qdd4_v,qdd5_v,qdd6_v = sess.run(dmpNet,feed_dict = feeddict)
            dx = [x[6],x[7],x[8],x[9],x[10],x[11],qdd1_v,qdd2_v,qdd3_v,qdd4_v,qdd5_v,qdd6_v]
            return dx

        period = 50
        t = np.linspace(2, period, 100)
        print(t)
        tau_v = float(1)/period
        # q0_v = [-0.0003235975848596695, -1.040771786366598, 1.6213598251342773, -0.34193402925600225, 1.5711277723312378, 3.141711950302124]
        q0_v = [0, -pi/2, pi/2, 0, pi/2, pi]
        v0 = [0,0,0,0,0,0]
        x0 = []; x0.extend(q0_v); x0.extend(v0)
        print("Initiate Object Detection")
        STATE = 'NONE'
        _,_,_, last_capture = IntelRealsense.pingpong_detection(display = False)

        T_camera2base = IntelRealsense.transformation_camera2base()
        T_end2base = IntelRealsense.transformation_end2base()
        gripper = np.array([  [0],
                              [0],
                              [0.1],
                              [1]  ])
        Y = np.array([ [0],
                       [0],
                       [0],
                       [0],
                       [0],
                       [0] ])
        X = np.array([ [np.random.normal(0, sigma_x)],
                       [np.random.normal(0, sigma_y)],
                       [np.random.normal(0, sigma_d)],
                       [np.random.normal(0, sigma_v_x)],
                       [np.random.normal(0, sigma_v_y)],
                       [np.random.normal(0, sigma_v_d)] ])
        X, P = init_state(X, Y, 0)
        print('First State is %s' %X)
        I = np.identity(6, dtype=float)
        H = np.identity(6, dtype=float)
        count = 0
        global client
        try:
            rospy.init_node("dmp", anonymous=True, disable_signals=True)
            client = actionlib.SimpleActionClient('follow_joint_trajectory', FollowJointTrajectoryAction)
            client.wait_for_server()
            print ("Connected to server")
            parameters = rospy.get_param(None)
            index = str(parameters).find('prefix')
            if (index > 0):
                prefix = str(parameters)[index+len("prefix': '"):(index+len("prefix': '")+str(parameters)[index+len("prefix': '"):-1].find("'"))]
                for i, name in enumerate(JOINT_NAMES):
                    JOINT_NAMES[i] = prefix + name
            book = Workbook()
            sheet = book.active
            excel_row = 1
            while(True):
                initial_time = time.time()
                pp_x, pp_y, pp_depth, capture = IntelRealsense.pingpong_detection(display = True)
                pingpong_camera = IntelRealsense.reverse_perspective_projection(pp_x, pp_y, pp_depth); pingpong_camera[3] = 1
                processing = capture - last_capture; timestep = processing
                pingpong_base = T_camera2base.dot(pingpong_camera)
                v_x, v_depth, v_z, a_x, a_y, a_depth, STATE = IntelRealsense.pingpong_velocity(STATE, pingpong_base[0,0], pingpong_base[1,0], pingpong_base[2,0], capture, display = False)
                last_capture = capture

                sheet.cell(row = excel_row,column = 1).value = pingpong_base[0]
                sheet.cell(row = excel_row,column = 2).value = pingpong_base[1]
                sheet.cell(row = excel_row,column = 3).value = pingpong_base[2]
                sheet.cell(row = excel_row,column = 4).value = v_x
                sheet.cell(row = excel_row,column = 5).value = v_depth
                sheet.cell(row = excel_row,column = 6).value = v_z
                sheet.cell(row = excel_row,column = 7).value = a_x
                sheet.cell(row = excel_row,column = 8).value = a_y
                sheet.cell(row = excel_row,column = 9).value = a_depth
                excel_row +=1
                if excel_row == 10000:
                    book.save('/home/s/catkin_ws/src/camera/src/real_world_variance.xlsx')
                    print('Excel Saved')
                # print('Real World Coordinates From KF \n %s \n' %pingpong_camera)
                # print('Ping-Pong Position (Base Frame) \n %s \n' %pingpong_base)
                grasping_point = [pingpong_base[0], pingpong_base[1], pingpong_base[2]]
                # print('grasping point %s' %grasping_point)
                pingpong_base = pingpong_base - T_end2base.dot(gripper)
                inp_invKine = IntelRealsense.transformation_end2base(pingpong_base[0,0],pingpong_base[1,0],pingpong_base[2,0])
                inv_pingpong = invKine(inp_invKine)
                # print('IK \n %s \n' %inv_pingpong)
                grasping_q = cost_func(inv_pingpong)
                # print('Grasping q \n%s\n' %grasping_q)
                g_v = grasping_q
                print('grasping_q %s' %grasping_q)
                fwd_grasping = fwd_kin(g_v, o_unit = 'p')
                print('fwd Grasping q %s' %fwd_grasping)
                inp = raw_input("Set Home! (Enter to continue)")
                set_home(set_duration = 10)
                # inp = raw_input("Continue? y/n: ")[0]
                # if (inp == 'y'):
                # q_ode = odeint(dynamics,x0,t,args=(tau_v,g_v,q0_v,grasping_point,sess))
                q_ode = np.load('q_sample1.npy')
                # set_home(g_v, 3)
                # print('q_ode %s' %q_ode)
                t = np.linspace(0.1, 10, 100)
                move_dmp_path(q_ode,t)
                final_time = time.time()
                print("Processing Time : %s" %(final_time-initial_time))
                # else:
                    # print ("Halting program")
                    # break
                # print('\n#############################################\n')
        except KeyboardInterrupt:
            rospy.signal_shutdown("KeyboardInterrupt")
            raise

if __name__ == '__main__': main()
